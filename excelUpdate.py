import openpyxl
import skills
import datetime


# insert the new cell in FRONT of the old cells.
def excel_update(keyword, prev_list, new_list):
    file_name = keyword + ".xlsx"
    book = openpyxl.load_workbook(file_name)
    sheet = book.active
    row = 3
    col = 1
    names_list = skills.skills()
    sheet.insert_cols(1, 7)
    date = "Date: " + (datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    sheet.cell(1, 1, 'Account Name-' + keyword)
    sheet.cell(2, 1, date)
    sheet.cell(2, 2, "Rank")
    sheet.cell(2, 3, "Level")
    sheet.cell(2, 4, "XP")
    sheet.cell(2, 5, "Rank Change")
    sheet.cell(2, 6, "XP Change")
    # add conditional formatting, if prev - new >0 green cell, <0 red cell

    # iterate through prev list and new list, pull rank and xp, and sub new from old
    for i in range(len(new_list)):
        sheet.cell(row, col, names_list[i])
        sheet.cell(row, col+1, new_list[i][0])
        sheet.cell(row, col+2, new_list[i][1])
        sheet.cell(row, col+3, new_list[i][2])
        sheet.cell(row, col+4, (int(prev_list[i][0]) - int(new_list[i][0])))
        sheet.cell(row, col+5, (int(new_list[i][2]) - int(prev_list[i][2])))
        row += 1

    book.save(file_name)
